import json
import io

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image

FULL_NAME = 'FullName'
NAME = 'Name'
URL = 'Url'
IMAGE_URL = 'ImageUrl'
IMAGE_PATH = 'ImagePath'
PRICE = 'Price'
LAUNCH_TIME = 'LaunchTime'


def read_json(path: str):
    with io.open(path, mode='r', encoding='utf-8') as fs:
        json_obj = json.load(fs)
        return json_obj


def resize_image(img: Image, max_height=320, max_width=200):
    cell_ratio = float(max_height) / max_width
    img_ratio = float(img.height) / img.width

    if cell_ratio < img_ratio:
        h_percent = max_height / float(img.height)
        w_size = int(float(img.width) * float(h_percent))
        img.height = max_height - 1
        img.width = w_size
    else:
        w_percent = max_width / float(img.width)
        h_size = int(float(img.height) * float(w_percent))
        img.height = h_size
        img.width = max_width - 1

    return img


def append_titles(ws: Worksheet):
    ws['A1'] = FULL_NAME
    ws['B1'] = NAME
    ws['C1'] = URL
    ws['D1'] = IMAGE_URL

    ws['E1'] = IMAGE_PATH
    ws.column_dimensions['E'].width = 20

    ws['F1'] = PRICE
    ws['G1'] = LAUNCH_TIME

    return ws


def append_cars(ws: Worksheet, cars):
    car_count = len(cars)
    for n in range(0, car_count):
        car = cars[n]

        row_index = n + 2
        ws['A' + str(row_index)] = car[FULL_NAME]
        ws['B' + str(row_index)] = car[NAME]
        ws['C' + str(row_index)] = car[URL]
        ws['D' + str(row_index)] = car[IMAGE_URL]

        car_img_path = car[IMAGE_PATH]
        if car_img_path != '':
            car_img = Image(car[IMAGE_PATH])
            car_img = resize_image(car_img)
            ws.add_image(car_img, anchor='E' + str(row_index))
            ws.row_dimensions[row_index].height = 108
        else:
            ws['E' + str(row_index)] = ''

        ws['F' + str(row_index)] = car[PRICE]
        ws['G' + str(row_index)] = car[LAUNCH_TIME]

    return ws


if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cars'
    ws = append_titles(ws)

    cars = read_json('data.json')
    ws = append_cars(ws, cars)

    wb.save('output.xlsx')
